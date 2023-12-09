from openpyxl import Workbook
from classes.state import State
from typing import List

# Your code that generates T, nT, and states

def filterForUniqueStates(states: List[State]):
    unique = [s for s in states if not s.isCopy]
    return unique

def exportLRTAble(T, nT, states):
    nT.remove('START')
    T.remove('$')

    # Create a new workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LR_Table"

    # Columns - symbols
    symbols = [''] + T + nT + ['#']
    for index, symbol in enumerate(symbols, start=1):
        sheet.cell(row=1, column=index, value=symbol)

    # Rows - states
    states = filterForUniqueStates(states)
    statesRowNo = {}

    for row_number, state in enumerate(states, start=2):
        stateNo = state.orderNumber
        sheet.cell(row=row_number, column=1, value=f'state {stateNo}')
        statesRowNo[stateNo] = row_number

    # Populate the table
    for state in states:
        row_index = statesRowNo[state.orderNumber]

        for action_key, action_value in state.actions.items():
            column_index = symbols.index(action_key) + 1
            sheet.cell(row=row_index, column=column_index, value=f'{action_value}') 

    # Save to a file
    workbook.save("LR_Table.xlsx")
